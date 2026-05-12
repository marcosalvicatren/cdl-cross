#!/usr/bin/env python3
"""
Prima Nota Paghe  —  GB Software / Wolters Kluwer
Streamlit app con motore a regole e persistenza via GitHub API
"""

import streamlit as st
import json, re, io, base64, calendar, requests
from datetime import datetime
from pathlib import Path
import pandas as pd

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    openpyxl = None

try:
    from lxml import etree
except ImportError:
    etree = None

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURAZIONE
# ═══════════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Prima Nota Paghe",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@300;400;500;600&family=IBM+Plex+Mono:wght@400;500&display=swap');

html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }

section[data-testid="stSidebar"] { background: #0f1923; }
section[data-testid="stSidebar"] * { color: #c8d8e8 !important; }
section[data-testid="stSidebar"] hr { border-color: #1e3048; }

h1 { color: #0f1923 !important; font-weight: 600 !important; letter-spacing: -0.5px; }
h2 { color: #0f1923 !important; font-weight: 500 !important; }
h3 { color: #1a3a5c !important; font-weight: 500 !important; font-size: 1rem !important; }

.stButton > button {
    background: #0f1923 !important; color: #fff !important;
    border: none !important; border-radius: 4px !important;
    font-family: 'IBM Plex Sans', sans-serif !important;
    font-weight: 500 !important; padding: 8px 20px !important;
}
.stButton > button:hover { background: #1a3a5c !important; }

.stDownloadButton > button {
    background: #1a5c3a !important; color: #fff !important;
    border: none !important; border-radius: 4px !important;
    font-family: 'IBM Plex Sans', sans-serif !important; font-weight: 500 !important;
}

.badge-ok   { background:#e6f4ea; color:#1a5c3a; padding:3px 10px; border-radius:12px; font-size:12px; font-weight:500; }
.badge-warn { background:#fff8e1; color:#7a5c00; padding:3px 10px; border-radius:12px; font-size:12px; font-weight:500; }
.badge-err  { background:#fdecea; color:#7a1a1a; padding:3px 10px; border-radius:12px; font-size:12px; font-weight:500; }

.section-header {
    border-left: 3px solid #0f1923; padding: 6px 0 6px 14px;
    margin: 24px 0 12px 0; font-size: 15px; font-weight: 600; color: #0f1923;
}
.rule-hint { font-size: 12px; color: #6b7a8d; font-family: 'IBM Plex Mono', monospace; }
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# REGOLE DEFAULT
# ═══════════════════════════════════════════════════════════════════════════════

# Ogni regola: contiene, non_contiene, conto, da, desc_xml, riga_extra_conto, riga_extra_da
# Le regole vengono applicate in ordine: vince la prima che corrisponde.
# contiene / non_contiene: lista di stringhe (case-insensitive)

REGOLE_BP_DEFAULT = [
    {"contiene": ["stipend"],           "non_contiene": [],             "conto": "61000",   "da": "D", "desc_xml": "Stipendi",                                "riga_extra_conto": "",        "riga_extra_da": ""},
    {"contiene": ["inps", "ditta"],     "non_contiene": ["dipendente"], "conto": "61100",   "da": "D", "desc_xml": "Contributi INPS c/ditta",                 "riga_extra_conto": "",        "riga_extra_da": ""},
    {"contiene": ["solidariet"],        "non_contiene": [],             "conto": "61100",   "da": "D", "desc_xml": "Solidarietà 10% da enti",                 "riga_extra_conto": "",        "riga_extra_da": ""},
    {"contiene": ["metasalute"],        "non_contiene": [],             "conto": "6112601", "da": "D", "desc_xml": "METASALUTE c/azienda",                    "riga_extra_conto": "45424",   "riga_extra_da": "A"},
    {"contiene": ["tfr", "fondo"],      "non_contiene": [],             "conto": "61200",   "da": "D", "desc_xml": "Quota TFR fondo prev. complementare",     "riga_extra_conto": "4523001", "riga_extra_da": "A"},
    {"contiene": ["trattamento integr"],"non_contiene": ["f24","compensare"],"conto": "42055","da": "D","desc_xml": "Trattamento integrativo DL 3/20",        "riga_extra_conto": "",        "riga_extra_da": ""},
    {"contiene": ["bonus fiscal"],      "non_contiene": ["f24","compensare"],"conto": "42055","da": "D","desc_xml": "Bonus fiscali erogati",                  "riga_extra_conto": "",        "riga_extra_da": ""},
    {"contiene": ["saldo", "dm10"],     "non_contiene": [],             "conto": "45200",   "da": "A", "desc_xml": "Saldo a versare DM10",                    "riga_extra_conto": "",        "riga_extra_da": ""},
    {"contiene": ["irpef","1001"],      "non_contiene": ["addizional","restit","regionale","comunale"],"conto": "45000","da": "A","desc_xml": "IRPEF 1001-1013","riga_extra_conto": "",       "riga_extra_da": ""},
    {"contiene": ["irpef","restituit"], "non_contiene": [],             "conto": "45000",   "da": "D", "desc_xml": "IRPEF restituita",                        "riga_extra_conto": "",        "riga_extra_da": ""},
    {"contiene": ["addizionale","region"],"non_contiene": [],           "conto": "45010",   "da": "A", "desc_xml": "Addizionale IRPEF regionale",             "riga_extra_conto": "",        "riga_extra_da": ""},
    {"contiene": ["addizionale","comun"],"non_contiene": [],            "conto": "45011",   "da": "A", "desc_xml": "Addizionale IRPEF comunale",              "riga_extra_conto": "",        "riga_extra_da": ""},
    {"contiene": ["netto in busta"],    "non_contiene": [],             "conto": "45420",   "da": "A", "desc_xml": "Netto in busta",                          "riga_extra_conto": "",        "riga_extra_da": ""},
    {"contiene": ["arrotondament"],     "non_contiene": [],             "conto": "70621",   "da": "",  "desc_xml": "Arrotondamenti",                          "riga_extra_conto": "",        "riga_extra_da": ""},
    {"contiene": ["rimborso spese"],    "non_contiene": [],             "conto": "60813",   "da": "D", "desc_xml": "Rimborso spese documentate",              "riga_extra_conto": "",        "riga_extra_da": ""},
    {"contiene": ["ulteriore detraz"],  "non_contiene": [],             "conto": "",        "da": "D", "desc_xml": "Ulteriore detrazione",                    "riga_extra_conto": "",        "riga_extra_da": ""},
]

REGOLE_F24_DEFAULT = [
    {"contiene": ["1001"],  "non_contiene": [],      "conto": "45000", "da": "D", "desc_xml": "IRPEF 1001-1013"},
    {"contiene": ["1627"],  "non_contiene": [],      "conto": "45000", "da": "A", "desc_xml": "Conguaglio IRPEF fine anno"},
    {"contiene": ["1701"],  "non_contiene": [],      "conto": "42055", "da": "A", "desc_xml": "Trattamento integrativo DL 3/20"},
    {"contiene": ["1704"],  "non_contiene": [],      "conto": "42055", "da": "A", "desc_xml": "Bonus fiscali erogati"},
    {"contiene": ["1712"],  "non_contiene": [],      "conto": "42143", "da": "D", "desc_xml": "Imposta sostitutiva TFR"},
    {"contiene": ["dm10"],  "non_contiene": [],      "conto": "45200", "da": "D", "desc_xml": "Saldo DM10"},
    {"contiene": ["met1"],  "non_contiene": [],      "conto": "45424", "da": "D", "desc_xml": "METASALUTE c/azienda"},
    {"contiene": ["3802"],  "non_contiene": [],      "conto": "45010", "da": "D", "desc_xml": "Addizionale IRPEF regionale"},
    {"contiene": ["3847"],  "non_contiene": ["3848"],"conto": "45011", "da": "D", "desc_xml": "Addizionale IRPEF comunale"},
    {"contiene": ["3848"],  "non_contiene": [],      "conto": "45011", "da": "D", "desc_xml": "Addizionale IRPEF comunale (AP)"},
]

ESCLUDI_BP = [
    "contributi inps c/dipendente",
    "contributi inail",
    "tratt. integrativo da compensare",
    "bonus fiscali da compensare",
]

SEZIONI_ESCLUDI = {"TRATTENUTE FISCALI (IRPEF)", "CREDITI IRPEF"}

# ═══════════════════════════════════════════════════════════════════════════════
# GITHUB PERSISTENCE
# ═══════════════════════════════════════════════════════════════════════════════

def _gh_headers():
    token = st.secrets.get("GITHUB_TOKEN", "")
    if not token:
        return None
    return {"Authorization": f"token {token}", "Accept": "application/vnd.github.v3+json"}

def _gh_coords():
    return (
        st.secrets.get("GITHUB_REPO", ""),   # es. "utente/prima-nota-paghe"
        st.secrets.get("GITHUB_BRANCH", "main"),
    )

def carica_regole_da_github(filename):
    """Carica un file JSON dal repo GitHub. Ritorna None se non configurato o non trovato."""
    headers = _gh_headers()
    repo, branch = _gh_coords()
    if not headers or not repo:
        return None
    url = f"https://api.github.com/repos/{repo}/contents/{filename}?ref={branch}"
    r = requests.get(url, headers=headers, timeout=10)
    if r.status_code != 200:
        return None
    data = r.json()
    content = base64.b64decode(data["content"]).decode("utf-8")
    return json.loads(content), data["sha"]

def salva_regole_su_github(filename, contenuto, messaggio_commit):
    """Salva / aggiorna un file JSON nel repo GitHub."""
    headers = _gh_headers()
    repo, branch = _gh_coords()
    if not headers or not repo:
        return False, "GitHub non configurato (vedi secrets)"

    # Recupera SHA attuale se il file esiste già
    sha = None
    url = f"https://api.github.com/repos/{repo}/contents/{filename}"
    r = requests.get(url, headers=headers, params={"ref": branch}, timeout=10)
    if r.status_code == 200:
        sha = r.json()["sha"]

    payload = {
        "message": messaggio_commit,
        "content": base64.b64encode(json.dumps(contenuto, ensure_ascii=False, indent=2).encode()).decode(),
        "branch": branch,
    }
    if sha:
        payload["sha"] = sha

    r = requests.put(url, headers=headers, json=payload, timeout=15)
    if r.status_code in (200, 201):
        return True, "Salvato su GitHub ✓"
    return False, f"Errore GitHub: {r.status_code} — {r.json().get('message','')}"

# ═══════════════════════════════════════════════════════════════════════════════
# SESSION STATE  —  carica regole all'avvio
# ═══════════════════════════════════════════════════════════════════════════════

def init_state():
    if "regole_bp" not in st.session_state:
        result = carica_regole_da_github("regole/buste_paga.json")
        st.session_state.regole_bp = result[0] if result else REGOLE_BP_DEFAULT[:]

    if "regole_f24" not in st.session_state:
        result = carica_regole_da_github("regole/f24.json")
        st.session_state.regole_f24 = result[0] if result else REGOLE_F24_DEFAULT[:]

    if "escludi_bp" not in st.session_state:
        st.session_state.escludi_bp = ESCLUDI_BP[:]

    for k in ("bp_voci","f24_voci","bp_mese","bp_azienda","f24_mese","bp_data","f24_data","f24_sconosciuti"):
        if k not in st.session_state:
            st.session_state[k] = None

init_state()

# ═══════════════════════════════════════════════════════════════════════════════
# MOTORE A REGOLE
# ═══════════════════════════════════════════════════════════════════════════════

def applica_regola(desc_lower, regola):
    """Ritorna True se la descrizione soddisfa la regola."""
    for c in regola.get("contiene", []):
        if c.lower() not in desc_lower:
            return False
    for nc in regola.get("non_contiene", []):
        if nc.lower() in desc_lower:
            return False
    return True

def classifica_voce_bp(desc_pdf, da_pdf, importo):
    """
    Applica le regole BP in ordine.
    Ritorna lista di dict (1 o 2 righe se voce doppia).
    Se nessuna regola corrisponde, ritorna voce con conto vuoto.
    """
    desc_lower = desc_pdf.lower()

    # Escludi voci da ignorare
    for escludi in st.session_state.escludi_bp:
        if escludi.lower() in desc_lower:
            return []  # voce scartata silenziosamente

    for regola in st.session_state.regole_bp:
        if applica_regola(desc_lower, regola):
            da = regola.get("da") or da_pdf or "D"
            righe = [{
                "desc_xml": regola["desc_xml"],
                "importo":  importo,
                "da":       da,
                "conto":    regola.get("conto", ""),
                "causale":  "LA",
                "regola":   " + ".join(regola.get("contiene", [])),
            }]
            # Riga extra (es. METASALUTE, TFR fondo)
            if regola.get("riga_extra_conto"):
                righe.append({
                    "desc_xml": regola["desc_xml"],
                    "importo":  importo,
                    "da":       regola["riga_extra_da"],
                    "conto":    regola["riga_extra_conto"],
                    "causale":  "LA",
                    "regola":   "riga automatica",
                })
            return righe

    # Nessuna regola: voce manuale con conto vuoto
    return [{
        "desc_xml": desc_pdf,
        "importo":  importo,
        "da":       da_pdf or "D",
        "conto":    "",
        "causale":  "LA",
        "regola":   "⚠ nessuna regola — compilare manualmente",
    }]

# Lookup diretto per codici F24 estratti dalla sezione INPS (sempre esatti)
_CODICI_FISSI_F24 = {
    "DM10": {"desc_xml": "Saldo DM10",          "conto": "45200", "da": "D"},
    "MET1": {"desc_xml": "METASALUTE c/azienda", "conto": "45424", "da": "D"},
    "INAIL":{"desc_xml": "INAIL credito",        "conto": "45300", "da": "A"},
}

def classifica_voce_f24(codice, importo):
    # 1. Corrispondenza esatta per codici INPS (DM10, MET1, INAIL)
    if codice.upper() in _CODICI_FISSI_F24:
        r = _CODICI_FISSI_F24[codice.upper()]
        return {"codice": codice, "desc_xml": r["desc_xml"],
                "importo": importo, "da": r["da"], "conto": r["conto"], "causale": "LA"}
    # 2. Regole configurabili per tutto il resto
    codice_lower = codice.lower()
    for regola in st.session_state.regole_f24:
        if applica_regola(codice_lower, regola):
            return {
                "codice":   codice,
                "desc_xml": regola["desc_xml"],
                "importo":  importo,
                "da":       regola.get("da", "D"),
                "conto":    regola.get("conto", ""),
                "causale":  "LA",
            }
    return None  # non mappato

# ═══════════════════════════════════════════════════════════════════════════════
# PARSER PDF  —  BUSTE PAGA
# ═══════════════════════════════════════════════════════════════════════════════

SEZIONI_LABEL = [
    "RETRIBUZIONI E ALTRE COMPETENZE", "CONTRIBUTI INPS",
    "ALTRI VERSAMENTI", "TRATTENUTE FISCALI (IRPEF)",
    "CREDITI IRPEF", "CREDITI E BONUS FISCALI",
]
DESC_SUBTOTALI = ["Totale altri versamenti","TOTALE irpef dovuta","Totale crediti Irpef"]

def parse_buste_paga(pdf_bytes, page_num):
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        if page_num < 1 or page_num > len(pdf.pages):
            raise ValueError(f"Pagina {page_num} non esiste (il PDF ha {len(pdf.pages)} pagine)")
        text = pdf.pages[page_num - 1].extract_text() or ""

    if not text.strip():
        raise ValueError(f"La pagina {page_num} è vuota o non leggibile")

    lines = text.split("\n")

    # Intestazione
    mese_anno = azienda = ""
    for l in lines[:10]:
        m = re.search(r'mese di\s+(\w+\s+\d{4})', l, re.IGNORECASE)
        if m: mese_anno = m.group(1).strip()
        if any(x in l for x in ["SRL","SPA","SNGL","srl","spa"]) and "Azienda" in l:
            parts = l.split()
            idx = next((i for i,p in enumerate(parts) if re.match(r'^\d{4}$', p)), 1)
            azienda = " ".join(parts[idx+1:])

    voci = []
    sezione = ""
    for line in lines:
        for s in SEZIONI_LABEL:
            if s in line: sezione = s; break

        if sezione in SEZIONI_ESCLUDI: continue
        if not line.strip().startswith("__________"): continue

        rest = line.strip()[10:].strip()
        m = re.search(r'([\d\.]+,\d{2})\s*([DA]?)$', rest)
        if not m: continue

        importo_str = m.group(1)
        da_pdf      = m.group(2)
        desc_pdf    = rest[:m.start()].strip()

        if any(sub in desc_pdf for sub in DESC_SUBTOTALI): continue

        importo = float(importo_str.replace(".", "").replace(",", "."))

        # Sezione bonus: salta righe "A" (doppioni)
        if sezione == "CREDITI E BONUS FISCALI" and da_pdf == "A": continue

        righe = classifica_voce_bp(desc_pdf, da_pdf, importo)
        voci.extend(righe)

    return mese_anno, azienda, voci


def fmt_data_fine_mese(mese_anno):
    mesi = {"gennaio":"01","febbraio":"02","marzo":"03","aprile":"04",
            "maggio":"05","giugno":"06","luglio":"07","agosto":"08",
            "settembre":"09","ottobre":"10","novembre":"11","dicembre":"12"}
    parts = mese_anno.lower().split()
    if len(parts) == 2:
        mm = mesi.get(parts[0], "12")
        yy = parts[1]
        last = calendar.monthrange(int(yy), int(mm))[1]
        return f"{last:02d}/{mm}/{yy}"
    return ""

# ═══════════════════════════════════════════════════════════════════════════════
# PARSER PDF  —  F24
# ═══════════════════════════════════════════════════════════════════════════════

def _centesimi(s):
    """Converte stringa intera di centesimi (es. '11285') in float (112.85)"""
    try: return int(s) / 100
    except: return None

def trova_pagina_f24(pdf_bytes):
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for i, page in enumerate(pdf.pages, 1):
            text = page.extract_text() or ""
            if re.search(r'\b(1001|1704|1701|DM10|MET1|3802|3847|3848)\b', text):
                if re.search(r'\b1500\s+DM10\b|\b\d{4}\s+\d{1,2}\s+20\d{2}\s+\d{3,6}\b', text):
                    return i
    return None

def parse_f24(pdf_bytes, page_num):
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        if page_num < 1 or page_num > len(pdf.pages):
            raise ValueError(f"Pagina {page_num} non esiste")
        text = pdf.pages[page_num-1].extract_text() or ""

    # Estrai mese/anno dalla riga INPS (es. "1500 DM10 1519416621 012026 92800")
    mese_anno = ""
    m = re.search(r'1500\s+(?:DM10|MET1)\s+\d+\s+(\d{2})(\d{4})', text)
    if m:
        mm, yy = int(m.group(1)), m.group(2)
        nomi = ["","gennaio","febbraio","marzo","aprile","maggio","giugno",
                "luglio","agosto","settembre","ottobre","novembre","dicembre"]
        if 1 <= mm <= 12:
            mese_anno = f"{nomi[mm].capitalize()} {yy}"

    righe_raw = []

    for line in text.split("\n"):
        line = line.strip()
        if not line: continue

        # SEZIONE ERARIO: CODICE mm AAAA importo_centesimi (solo debiti — no righe con +)
        m = re.match(r'^(\d{4})\s+(\d{1,2})\s+(20\d{2})\s+(\d+)$', line)
        if m:
            imp = _centesimi(m.group(4))
            if imp: righe_raw.append({"codice": m.group(1), "importo": imp}); continue

        # SEZIONE INPS: 1500 DM10/MET1 matricola mmAAAA importo_centesimi
        m = re.match(r'^1500\s+(DM10|MET1)\s+\d+\s+\d{6}\s+(\d+)$', line)
        if m:
            imp = _centesimi(m.group(2))
            if imp: righe_raw.append({"codice": m.group(1), "importo": imp}); continue

        # SEZIONE REGIONI: codice_regione CODICE mm AAAA importo_centesimi
        m = re.match(r'^\d+\s+(\d{4})\s+(\d{1,2})\s+(20\d{2})\s+(\d+)$', line)
        if m:
            imp = _centesimi(m.group(4))
            if imp: righe_raw.append({"codice": m.group(1), "importo": imp}); continue

        # SEZIONE IMU/LOCALI: C293 CODICE mm AAAA importo_centesimi
        m = re.match(r'^[A-Z]\d+\s+(\d{4})\s+(\d{1,2})\s+(20\d{2})\s+(\d+)$', line)
        if m:
            imp = _centesimi(m.group(4))
            if imp: righe_raw.append({"codice": m.group(1), "importo": imp}); continue

        # INAIL credito: 13200 codice c.c. periodo causale importo_centesimi
        m = re.match(r'^13200\s+\S+\s+\d+\s+\S+\s+[A-Z]\s+(\d+)$', line)
        if m:
            imp = _centesimi(m.group(1))
            if imp: righe_raw.append({"codice": "INAIL", "importo": imp}); continue

        # Crediti erario: CODICE AAAA importo_centesimi (es. 6781 2025 997 = imp.sost.TFR credito)
        m = re.match(r'^(\d{4})\s+(20\d{2})\s+(\d+)$', line)
        if m:
            imp = _centesimi(m.group(3))
            if imp: righe_raw.append({"codice": m.group(1), "importo": imp, "credito": True}); continue

    voci = []; sconosciuti = []
    for r in righe_raw:
        v = classifica_voce_f24(r["codice"], r["importo"])
        # Per crediti (es. 6781 imp.sost.TFR) inverti il segno
        if v and r.get("credito"):
            v["da"] = "A" if v["da"] == "D" else "D"
        if v: voci.append(v)
        else: sconosciuti.append(r)

    # Data pagamento = 16 del mese successivo
    data_pag = ""
    mesi_map = {"gennaio":"01","febbraio":"02","marzo":"03","aprile":"04",
                "maggio":"05","giugno":"06","luglio":"07","agosto":"08",
                "settembre":"09","ottobre":"10","novembre":"11","dicembre":"12"}
    parts = mese_anno.lower().split()
    if len(parts) == 2:
        mm = int(mesi_map.get(parts[0], "1"))
        yy = int(parts[1])
        mm, yy = (1, yy+1) if mm == 12 else (mm+1, yy)
        data_pag = f"16/{mm:02d}/{yy}"

    return mese_anno, voci, sconosciuti, data_pag

# ═══════════════════════════════════════════════════════════════════════════════
# GENERATORI XLSX
# ═══════════════════════════════════════════════════════════════════════════════

def _brd():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)

def genera_xlsx_bp(voci, mese_anno, azienda):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Prima Nota Paghe"
    brd = _brd()
    blu = PatternFill("solid", fgColor="0F1923")
    hft = Font(bold=True, color="FFFFFF", name="Calibri", size=10)

    ws["A1"] = f"PRIMA NOTA PAGHE  —  {azienda}  —  {mese_anno.upper()}"
    ws["A1"].font = Font(bold=True, size=13, color="0F1923", name="Calibri")
    ws.merge_cells("A1:F1"); ws.row_dimensions[1].height = 28

    ws["A2"] = ("Descrizione: testo nell'XML  |  D/A: D=Dare A=Avere  |  "
                "Conto: precompilato dalle regole  |  Celle gialle = conto mancante, completare prima di generare l'XML")
    ws["A2"].font = Font(name="Calibri", size=8, italic=True, color="666666")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:F2"); ws.row_dimensions[2].height = 30

    hdrs   = ["Descrizione", "Importo €", "D/A", "Conto", "Causale", "Da importare"]
    widths = [52, 14, 6, 14, 10, 12]
    for col, (h, w) in enumerate(zip(hdrs, widths), 1):
        c = ws.cell(row=3, column=col, value=h)
        c.fill = blu; c.font = hft; c.border = brd
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 22
    ws.freeze_panes = "A4"

    alt1 = PatternFill("solid", fgColor="F4F7FB")
    alt2 = PatternFill("solid", fgColor="FFFFFF")
    giallo = PatternFill("solid", fgColor="FFF9C4")

    for i, v in enumerate(voci, 4):
        fill = alt1 if i % 2 == 0 else alt2
        manca_conto = not v.get("conto", "").strip()
        vals = [v["desc_xml"], v["importo"], v["da"], v.get("conto",""), v.get("causale","LA"), "SI"]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = Font(name="Calibri", size=9); c.border = brd
            c.fill = giallo if (col == 4 and manca_conto) else fill
            if col == 2:
                c.number_format = "#,##0.00"; c.alignment = Alignment(horizontal="right")
            elif col == 3:
                c.alignment = Alignment(horizontal="center")
                c.font = Font(name="Calibri", size=9, bold=True,
                              color="003399" if val == "D" else "990000")

    tr = len(voci) + 4
    tf = PatternFill("solid", fgColor="0F1923")
    for col in range(1, 7):
        ws.cell(row=tr, column=col).fill = tf; ws.cell(row=tr, column=col).border = brd
    ws.cell(row=tr, column=1, value="TOTALE").font = Font(bold=True, color="FFFFFF", name="Calibri")
    f = ws.cell(row=tr, column=2,
        value=f'=SUMIF(C4:C{tr-1},"D",B4:B{tr-1})-SUMIF(C4:C{tr-1},"A",B4:B{tr-1})')
    f.number_format = "#,##0.00"; f.font = Font(bold=True, color="FFFFFF", name="Calibri")

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def genera_xlsx_f24(voci, mese_anno, ctrop):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Prima Nota F24"
    brd = _brd()
    blu = PatternFill("solid", fgColor="0F1923")
    hft = Font(bold=True, color="FFFFFF", name="Calibri", size=10)

    ws["A1"] = f"PRIMA NOTA F24  —  {mese_anno.upper()}"
    ws["A1"].font = Font(bold=True, size=13, color="0F1923", name="Calibri")
    ws.merge_cells("A1:G1"); ws.row_dimensions[1].height = 28

    hdrs   = ["Cod. F24", "Descrizione", "Importo €", "D/A", "Conto", "Contropartita", "Causale"]
    widths = [10, 48, 14, 6, 12, 14, 10]
    for col, (h, w) in enumerate(zip(hdrs, widths), 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = blu; c.font = hft; c.border = brd
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 22; ws.freeze_panes = "A3"

    alt1 = PatternFill("solid", fgColor="F4F7FB"); alt2 = PatternFill("solid", fgColor="FFFFFF")
    for i, v in enumerate(voci, 3):
        fill = alt1 if i % 2 == 0 else alt2
        vals = [v.get("codice",""), v["desc_xml"], v["importo"], v["da"], v["conto"], ctrop, v.get("causale","LA")]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = Font(name="Calibri", size=9); c.border = brd; c.fill = fill
            if col == 3:
                c.number_format = "#,##0.00"; c.alignment = Alignment(horizontal="right")
            elif col == 4:
                c.alignment = Alignment(horizontal="center")
                c.font = Font(name="Calibri", size=9, bold=True,
                              color="003399" if val == "D" else "990000")

    tr = len(voci) + 3
    tf = PatternFill("solid", fgColor="0F1923")
    for col in range(1, 8):
        ws.cell(row=tr, column=col).fill = tf; ws.cell(row=tr, column=col).border = brd
    ws.cell(row=tr, column=2, value="TOTALE F24").font = Font(bold=True, color="FFFFFF", name="Calibri")
    tot = ws.cell(row=tr, column=3,
        value=f'=SUMIF(D3:D{tr-1},"D",C3:C{tr-1})-SUMIF(D3:D{tr-1},"A",C3:C{tr-1})')
    tot.number_format = "#,##0.00"; tot.font = Font(bold=True, color="FFFFFF", name="Calibri")

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

# ═══════════════════════════════════════════════════════════════════════════════
# LETTURA XLSX REVISIONATO
# ═══════════════════════════════════════════════════════════════════════════════

def leggi_xlsx_bp(xlsx_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True); ws = wb.active
    voci = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        desc = str(row[0] or "").strip()
        if not desc or desc == "TOTALE": continue
        if str(row[5] or "").upper() == "NO": continue
        try: imp = float(row[1])
        except: continue
        voci.append({"desc_xml": desc, "importo": imp, "da": str(row[2] or "").strip(),
                     "conto": str(row[3] or "").strip(), "causale": str(row[4] or "LA").strip()})
    return voci

def leggi_xlsx_f24(xlsx_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True); ws = wb.active
    voci = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        desc = str(row[1] or "").strip()
        if not desc or desc == "TOTALE F24": continue
        try: imp = float(row[2])
        except: continue
        voci.append({"codice": str(row[0] or ""), "desc_xml": desc, "importo": imp,
                     "da": str(row[3] or "").strip(), "conto": str(row[4] or "").strip(),
                     "contropartita": str(row[5] or "").strip(), "causale": str(row[6] or "LA").strip()})
    return voci

# ═══════════════════════════════════════════════════════════════════════════════
# GENERATORI XML
# ═══════════════════════════════════════════════════════════════════════════════

def _fmt_imp(v):
    return str(int(v)) if v == int(v) else f"{v:.2f}"

def _sub(p, tag, txt):
    e = etree.SubElement(p, tag); e.text = str(txt) if txt else ""; return e

def _fix_attrs(xml_str):
    """GB Software vuole noNamespaceSchemaLocation prima di xmlns:xsi."""
    def _swap(m):
        tag = m.group(0)
        xl = re.search(r'xmlns:xsi="[^"]*"', tag).group()
        ns = re.search(r'xsi:noNamespaceSchemaLocation="[^"]*"', tag).group()
        return f'<PrimaNotaXsd {ns} {xl}>'
    return re.sub(r'<PrimaNotaXsd[^>]+>', _swap, xml_str, count=1)

def genera_xml_bp(voci, data_iso, causale="LA", xsd="SchemaImportazionePrimaNotaV2.xsd"):
    XSI = "http://www.w3.org/2001/XMLSchema-instance"
    root = etree.Element("PrimaNotaXsd", nsmap={"xsi": XSI})
    root.set(f"{{{XSI}}}noNamespaceSchemaLocation", xsd)
    ni = etree.SubElement(etree.SubElement(etree.SubElement(
         root, "ListaPrimaNota"), "PrimaNotaImportazione"), "PrimaNotaNonIva")
    dg = etree.SubElement(ni, "PrimaNotaDatiGenerici")
    _sub(dg, "CausaleContabile",  causale)
    _sub(dg, "NumeroDocumento",   f"BP-{data_iso.replace('-','')}")
    _sub(dg, "DataDocumento",     data_iso)
    _sub(dg, "DataRegistrazione", data_iso)
    ld = etree.SubElement(etree.SubElement(ni, "PrimaNotaSezioneConto"), "ListaDettaglioSezioneConto")
    errori = []
    for i, v in enumerate(voci, 1):
        conto = v.get("conto","").strip()
        if not conto: errori.append(f"Riga {i} — '{v['desc_xml'][:40]}': conto mancante")
        iv = v["importo"] if v["da"] == "D" else -v["importo"]
        d = etree.SubElement(ld, "SezioneContoDettaglioNonIva")
        _sub(d, "Conto", conto); _sub(d, "ImponibileConto", _fmt_imp(iv)); _sub(d, "Descrizione", v["desc_xml"][:255])
    xml_str = _fix_attrs(etree.tostring(root, xml_declaration=True, encoding="UTF-8", pretty_print=True).decode())
    return xml_str.encode("UTF-8"), errori

def genera_xml_f24(voci, data_iso, ctrop="4420901", xsd="SchemaImportazionePrimaNotaV2.xsd"):
    XSI = "http://www.w3.org/2001/XMLSchema-instance"
    root = etree.Element("PrimaNotaXsd", nsmap={"xsi": XSI})
    root.set(f"{{{XSI}}}noNamespaceSchemaLocation", xsd)
    ni = etree.SubElement(etree.SubElement(etree.SubElement(
         root, "ListaPrimaNota"), "PrimaNotaImportazione"), "PrimaNotaNonIva")
    dg = etree.SubElement(ni, "PrimaNotaDatiGenerici")
    _sub(dg, "CausaleContabile",  "LA")
    _sub(dg, "NumeroDocumento",   f"F24-{data_iso.replace('-','')}")
    _sub(dg, "DataDocumento",     data_iso)
    _sub(dg, "DataRegistrazione", data_iso)
    ld = etree.SubElement(etree.SubElement(ni, "PrimaNotaSezioneConto"), "ListaDettaglioSezioneConto")
    errori = []
    for i, v in enumerate(voci, 1):
        conto = v.get("conto","").strip()
        if not conto: errori.append(f"Riga {i}: conto mancante")
        iv = v["importo"] if v["da"] == "D" else -v["importo"]
        d = etree.SubElement(ld, "SezioneContoDettaglioNonIva")
        _sub(d, "Conto", conto); _sub(d, "ImponibileConto", _fmt_imp(iv)); _sub(d, "Descrizione", v["desc_xml"][:255])
    td = sum(v["importo"] for v in voci if v["da"]=="D")
    ta = sum(v["importo"] for v in voci if v["da"]=="A")
    d = etree.SubElement(ld, "SezioneContoDettaglioNonIva")
    _sub(d, "Conto", ctrop); _sub(d, "ImponibileConto", _fmt_imp(-(td-ta))); _sub(d, "Descrizione", "Pagamento F24")
    xml_str = _fix_attrs(etree.tostring(root, xml_declaration=True, encoding="UTF-8", pretty_print=True).decode())
    return xml_str.encode("UTF-8"), errori

# ═══════════════════════════════════════════════════════════════════════════════
# UI HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def sezione(titolo):
    st.markdown(f'<div class="section-header">{titolo}</div>', unsafe_allow_html=True)

def badge(testo, tipo="ok"):
    st.markdown(f'<span class="badge-{tipo}">{testo}</span>', unsafe_allow_html=True)

def mostra_voci(voci, chiave_conto="conto"):
    if not voci: return
    rows = []
    for v in voci:
        manca = not v.get(chiave_conto,"").strip()
        rows.append({
            "Descrizione":  v.get("desc_xml", v.get("descrizione","")),
            "Importo €":    f"{v['importo']:,.2f}",
            "D/A":          v.get("da",""),
            "Conto":        v.get(chiave_conto,"") or "⚠ mancante",
        })
    df = pd.DataFrame(rows)
    st.dataframe(df, use_container_width=True, hide_index=True)

    td = sum(v["importo"] for v in voci if v.get("da")=="D")
    ta = sum(v["importo"] for v in voci if v.get("da")=="A")
    c1,c2,c3 = st.columns(3)
    c1.metric("Totale Dare",  f"€ {td:,.2f}")
    c2.metric("Totale Avere", f"€ {ta:,.2f}")
    c3.metric("Sbilancio",    f"€ {abs(td-ta):,.2f}")

# ═══════════════════════════════════════════════════════════════════════════════
# PAGINA — BUSTE PAGA
# ═══════════════════════════════════════════════════════════════════════════════

def pagina_buste_paga():
    st.title("📋 Buste Paga → Prima Nota")

    # ── Passo 1 ───────────────────────────────────────────────────────────────
    sezione("Passo 1 — Carica il PDF del riepilogo paghe")
    pdf_file = st.file_uploader("PDF riepilogo paghe", type=["pdf"], key="bp_pdf",
                                 label_visibility="collapsed")
    if pdf_file:
        c1, c2 = st.columns([2,1])
        with c1:
            page_num = st.number_input("Numero di pagina del riepilogo",
                                        min_value=1, max_value=50, value=9,
                                        help="Di solito è pagina 7 o 9 — aprire il PDF per verificare")
        with c2:
            st.write(""); st.write("")
            avvia = st.button("▶  Estrai voci", use_container_width=True)

        if avvia:
            if not pdfplumber:
                st.error("Libreria pdfplumber non installata.")
            else:
                with st.spinner("Lettura PDF..."):
                    try:
                        mese, az, voci = parse_buste_paga(pdf_file.read(), page_num)
                        st.session_state.bp_voci    = voci
                        st.session_state.bp_mese    = mese
                        st.session_state.bp_azienda = az
                        st.session_state.bp_data    = fmt_data_fine_mese(mese)
                    except Exception as e:
                        st.error(f"Errore: {e}")

    # ── Passo 2 ───────────────────────────────────────────────────────────────
    if st.session_state.bp_voci:
        voci  = st.session_state.bp_voci
        mese  = st.session_state.bp_mese or ""
        az    = st.session_state.bp_azienda or ""

        sezione(f"Passo 2 — Verifica e modifica le voci  ·  {az}  ·  {mese.upper()}")

        non_mapp = [v for v in voci if not v.get("conto","").strip()]
        if non_mapp:
            st.warning(f"⚠ {len(non_mapp)} voci senza conto — completale prima di generare l'XML. "
                       "Scarica l'XLSX, compila i campi evidenziati in giallo e ricaricalo al Passo 3.")
        else:
            st.success(f"✅ {len(voci)} voci estratte — tutti i conti sono precompilati dalle regole.")

        mostra_voci(voci)

        st.write("")
        xlsx_bytes = genera_xlsx_bp(voci, mese, az)
        nome_xlsx  = f"prima_nota_paghe_{mese.replace(' ','_')}.xlsx"
        st.download_button("⬇  Scarica XLSX", data=xlsx_bytes, file_name=nome_xlsx,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)

        # ── Passo 3 ───────────────────────────────────────────────────────────
        sezione("Passo 3 — Genera XML per GB Software")

        if non_mapp:
            st.info("Se hai completato i conti mancanti nell'XLSX, ricaricalo qui sotto. "
                    "Altrimenti puoi procedere e completare i conti direttamente nei campi.")

        xlsx_rev = st.file_uploader("Ricarica XLSX revisionato (opzionale)",
                                     type=["xlsx"], key="bp_xlsx_rev",
                                     label_visibility="collapsed")
        if xlsx_rev:
            try:
                voci_xml = leggi_xlsx_bp(xlsx_rev.read())
                st.success(f"✅ XLSX caricato: {len(voci_xml)} voci.")
            except Exception as e:
                st.error(f"Errore XLSX: {e}"); voci_xml = voci
        else:
            voci_xml = voci

        c1,c2,c3 = st.columns(3)
        with c1:
            data_doc = st.text_input("Data documento (gg/mm/aaaa)",
                                      value=st.session_state.bp_data or "",
                                      placeholder="es. 31/10/2025")
        with c2:
            causale = st.text_input("Causale contabile", value="LA")
        with c3:
            xsd = st.text_input("File XSD", value="SchemaImportazionePrimaNotaV2.xsd")

        if st.button("⚡  Genera XML", use_container_width=True):
            try:
                data_iso = datetime.strptime(data_doc.strip(), "%d/%m/%Y").strftime("%Y-%m-%d")
            except:
                st.error("Data non valida — formato atteso: gg/mm/aaaa"); st.stop()

            xml_bytes, errori = genera_xml_bp(voci_xml, data_iso, causale, xsd)
            nome_xml = f"prima_nota_paghe_{mese.replace(' ','_')}.xml"

            if errori:
                st.warning("⚠ Conti mancanti nell'XML:\n" + "\n".join(errori))
            else:
                st.success(f"✅ XML generato — {len(voci_xml)} righe.")

            st.download_button("⬇  Scarica XML", data=xml_bytes, file_name=nome_xml,
                               mime="application/xml", use_container_width=True)
            with st.expander("Anteprima XML"):
                st.code(xml_bytes.decode("UTF-8")[:3000], language="xml")

# ═══════════════════════════════════════════════════════════════════════════════
# PAGINA — F24
# ═══════════════════════════════════════════════════════════════════════════════

def pagina_f24():
    st.title("🏦 F24 → Prima Nota")

    CONTROPARTITE = {
        "4420901 — Debiti per ritenute da versare": "4420901",
        "5010001 — Banca c/c":                     "5010001",
    }

    sezione("Passo 1 — Carica il PDF delle buste paga")
    pdf_file = st.file_uploader("PDF buste paga", type=["pdf"], key="f24_pdf",
                                 label_visibility="collapsed")

    if pdf_file:
        c1,c2,c3 = st.columns(3)
        with c1:
            page_num = st.number_input("Pagina F24", min_value=1, max_value=99, value=3)
        with c2:
            ctrop_label = st.selectbox("Conto contropartita", list(CONTROPARTITE.keys()))
        with c3:
            st.write(""); st.write("")
            cerca = st.button("🔍 Cerca pagina automaticamente")

        if cerca:
            with st.spinner("Ricerca..."):
                p = trova_pagina_f24(pdf_file.read())
                pdf_file.seek(0)
            if p:
                st.success(f"✅ Pagina F24 trovata: pagina {p}")
            else:
                st.warning("Pagina non trovata — impostare manualmente")

        if st.button("▶  Estrai voci F24", use_container_width=True):
            if not pdfplumber:
                st.error("pdfplumber non installato.")
            else:
                with st.spinner("Lettura PDF..."):
                    try:
                        mese, voci, scon, data_pag = parse_f24(pdf_file.read(), page_num)
                        st.session_state.f24_voci        = voci
                        st.session_state.f24_mese        = mese
                        st.session_state.f24_sconosciuti = scon
                        st.session_state.f24_data        = data_pag
                        st.session_state.f24_ctrop       = CONTROPARTITE[ctrop_label]
                    except Exception as e:
                        st.error(f"Errore: {e}")

    if st.session_state.f24_voci:
        voci  = st.session_state.f24_voci
        mese  = st.session_state.f24_mese or ""
        ctrop = st.session_state.f24_ctrop or "4420901"
        scon  = st.session_state.f24_sconosciuti or []

        sezione(f"Passo 2 — Verifica voci  ·  {mese.upper()}")

        if scon:
            st.warning(f"⚠ {len(scon)} codici non riconosciuti dalle regole "
                       f"({', '.join(s['codice'] for s in scon)}) — "
                       "non inclusi nell'XML. Se si ripeteranno, aggiungerli nelle regole F24.")
        if voci:
            st.success(f"✅ {len(voci)} voci estratte.")

        mostra_voci(voci)

        xlsx_bytes = genera_xlsx_f24(voci, mese, ctrop)
        st.download_button("⬇  Scarica XLSX", data=xlsx_bytes,
                           file_name=f"prima_nota_f24_{mese.replace(' ','_')}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)

        sezione("Passo 3 — Genera XML")

        xlsx_rev = st.file_uploader("Ricarica XLSX revisionato (opzionale)",
                                     type=["xlsx"], key="f24_xlsx_rev",
                                     label_visibility="collapsed")
        if xlsx_rev:
            try:
                voci_xml = leggi_xlsx_f24(xlsx_rev.read())
                st.success(f"✅ XLSX caricato: {len(voci_xml)} voci.")
            except Exception as e:
                st.error(f"Errore XLSX: {e}"); voci_xml = voci
        else:
            voci_xml = voci

        c1,c2 = st.columns(2)
        with c1:
            data_doc = st.text_input("Data pagamento F24 (gg/mm/aaaa)",
                                      value=st.session_state.f24_data or "",
                                      placeholder="es. 16/11/2025")
        with c2:
            xsd = st.text_input("File XSD", value="SchemaImportazionePrimaNotaV2.xsd", key="f24_xsd")

        if st.button("⚡  Genera XML", use_container_width=True, key="f24_xml"):
            try:
                data_iso = datetime.strptime(data_doc.strip(), "%d/%m/%Y").strftime("%Y-%m-%d")
            except:
                st.error("Data non valida — formato: gg/mm/aaaa"); st.stop()

            xml_bytes, errori = genera_xml_f24(voci_xml, data_iso, ctrop, xsd)
            nome_xml = f"prima_nota_f24_{mese.replace(' ','_')}.xml"
            if errori:
                st.warning("⚠ " + "\n".join(errori))
            else:
                td = sum(v["importo"] for v in voci_xml if v["da"]=="D")
                ta = sum(v["importo"] for v in voci_xml if v["da"]=="A")
                st.success(f"✅ XML generato — {len(voci_xml)} voci — Netto F24: € {td-ta:,.2f}")

            st.download_button("⬇  Scarica XML", data=xml_bytes, file_name=nome_xml,
                               mime="application/xml", use_container_width=True)
            with st.expander("Anteprima XML"):
                st.code(xml_bytes.decode("UTF-8")[:3000], language="xml")

# ═══════════════════════════════════════════════════════════════════════════════
# PAGINA — REGOLE
# ═══════════════════════════════════════════════════════════════════════════════

def pagina_regole():
    st.title("⚙️ Regole di classificazione")
    st.markdown(
        "Le regole determinano come ogni voce del PDF viene associata a un conto contabile. "
        "Vengono applicate **in ordine**: vince la prima che corrisponde. "
        "Se nessuna regola corrisponde, la voce compare con il conto vuoto — compilabile a mano senza toccare le regole."
    )

    gh_ok = bool(st.secrets.get("GITHUB_TOKEN","") and st.secrets.get("GITHUB_REPO",""))
    if not gh_ok:
        st.info("ℹ️  GitHub non configurato nei secrets — le modifiche valgono solo per questa sessione. "
                "Configura GITHUB_TOKEN e GITHUB_REPO per salvarle in modo permanente.")

    tab_bp, tab_f24 = st.tabs(["📋  Buste Paga", "🏦  F24"])

    # ── Buste Paga ────────────────────────────────────────────────────────────
    with tab_bp:
        st.markdown('<p class="rule-hint">contiene / non_contiene: parole chiave separate da virgola (case-insensitive) · riga_extra_conto: lasciare vuoto se non serve</p>', unsafe_allow_html=True)

        df_bp = pd.DataFrame(st.session_state.regole_bp)
        # Assicura colonne
        for col in ["contiene","non_contiene","conto","da","desc_xml","riga_extra_conto","riga_extra_da"]:
            if col not in df_bp.columns: df_bp[col] = ""
        # Liste → stringa per l'editor
        df_bp["contiene"]     = df_bp["contiene"].apply(lambda x: ", ".join(x) if isinstance(x,list) else x)
        df_bp["non_contiene"] = df_bp["non_contiene"].apply(lambda x: ", ".join(x) if isinstance(x,list) else x)

        edited = st.data_editor(
            df_bp,
            column_config={
                "contiene":         st.column_config.TextColumn("Contiene (parole chiave)",     width="medium"),
                "non_contiene":     st.column_config.TextColumn("NON contiene",                 width="medium"),
                "desc_xml":         st.column_config.TextColumn("Descrizione nell'XML",         width="large"),
                "conto":            st.column_config.TextColumn("Conto",                        width="small"),
                "da":               st.column_config.SelectboxColumn("D/A", options=["D","A",""], width="small"),
                "riga_extra_conto": st.column_config.TextColumn("Conto riga 2 (opz.)",          width="small"),
                "riga_extra_da":    st.column_config.SelectboxColumn("D/A riga 2", options=["D","A",""], width="small"),
            },
            num_rows="dynamic",
            use_container_width=True,
            key="edit_bp"
        )

        c1,c2 = st.columns(2)
        with c1:
            if st.button("💾  Salva regole Buste Paga", use_container_width=True):
                # Riconverti stringhe → liste
                nuove = edited.to_dict("records")
                for r in nuove:
                    r["contiene"]     = [x.strip() for x in str(r.get("contiene","")).split(",") if x.strip()]
                    r["non_contiene"] = [x.strip() for x in str(r.get("non_contiene","")).split(",") if x.strip()]
                st.session_state.regole_bp = nuove
                if gh_ok:
                    ok, msg = salva_regole_su_github("regole/buste_paga.json", nuove,
                                                      f"Aggiornamento regole buste paga {datetime.now().strftime('%d/%m/%Y %H:%M')}")
                    st.success(msg) if ok else st.error(msg)
                else:
                    st.success("Regole aggiornate per questa sessione.")
        with c2:
            st.download_button("⬇  Esporta regole",
                               data=json.dumps(st.session_state.regole_bp, ensure_ascii=False, indent=2).encode(),
                               file_name="regole_buste_paga.json", mime="application/json",
                               use_container_width=True)

        imp = st.file_uploader("⬆  Importa regole (JSON)", type=["json"], key="imp_regole_bp",
                                label_visibility="collapsed")
        if imp:
            try:
                nuove = json.load(imp)
                st.session_state.regole_bp = nuove
                if gh_ok:
                    ok, msg = salva_regole_su_github("regole/buste_paga.json", nuove,
                                                      f"Import regole buste paga {datetime.now().strftime('%d/%m/%Y %H:%M')}")
                    st.success(msg) if ok else st.error(msg)
                else:
                    st.success("Regole importate per questa sessione.")
                st.rerun()
            except Exception as e:
                st.error(f"File non valido: {e}")

    # ── F24 ───────────────────────────────────────────────────────────────────
    with tab_f24:
        st.markdown('<p class="rule-hint">Per F24 le parole chiave corrispondono al codice tributo (es. "1001", "dm10")</p>', unsafe_allow_html=True)

        df_f24 = pd.DataFrame(st.session_state.regole_f24)
        for col in ["contiene","non_contiene","conto","da","desc_xml"]:
            if col not in df_f24.columns: df_f24[col] = ""
        df_f24["contiene"]     = df_f24["contiene"].apply(lambda x: ", ".join(x) if isinstance(x,list) else x)
        df_f24["non_contiene"] = df_f24["non_contiene"].apply(lambda x: ", ".join(x) if isinstance(x,list) else x)

        edited_f24 = st.data_editor(
            df_f24,
            column_config={
                "contiene":     st.column_config.TextColumn("Codice / parola chiave",  width="medium"),
                "non_contiene": st.column_config.TextColumn("NON contiene",            width="medium"),
                "desc_xml":     st.column_config.TextColumn("Descrizione nell'XML",    width="large"),
                "conto":        st.column_config.TextColumn("Conto",                   width="small"),
                "da":           st.column_config.SelectboxColumn("D/A", options=["D","A"], width="small"),
            },
            num_rows="dynamic",
            use_container_width=True,
            key="edit_f24"
        )

        c1,c2 = st.columns(2)
        with c1:
            if st.button("💾  Salva regole F24", use_container_width=True):
                nuove = edited_f24.to_dict("records")
                for r in nuove:
                    r["contiene"]     = [x.strip() for x in str(r.get("contiene","")).split(",") if x.strip()]
                    r["non_contiene"] = [x.strip() for x in str(r.get("non_contiene","")).split(",") if x.strip()]
                st.session_state.regole_f24 = nuove
                if gh_ok:
                    ok, msg = salva_regole_su_github("regole/f24.json", nuove,
                                                      f"Aggiornamento regole F24 {datetime.now().strftime('%d/%m/%Y %H:%M')}")
                    st.success(msg) if ok else st.error(msg)
                else:
                    st.success("Regole aggiornate per questa sessione.")
        with c2:
            st.download_button("⬇  Esporta regole",
                               data=json.dumps(st.session_state.regole_f24, ensure_ascii=False, indent=2).encode(),
                               file_name="regole_f24.json", mime="application/json",
                               use_container_width=True)

        imp2 = st.file_uploader("⬆  Importa regole (JSON)", type=["json"], key="imp_regole_f24",
                                 label_visibility="collapsed")
        if imp2:
            try:
                nuove = json.load(imp2)
                st.session_state.regole_f24 = nuove
                if gh_ok:
                    ok, msg = salva_regole_su_github("regole/f24.json", nuove,
                                                      f"Import regole F24 {datetime.now().strftime('%d/%m/%Y %H:%M')}")
                    st.success(msg) if ok else st.error(msg)
                else:
                    st.success("Regole importate per questa sessione.")
                st.rerun()
            except Exception as e:
                st.error(f"File non valido: {e}")

# ═══════════════════════════════════════════════════════════════════════════════
# SIDEBAR + ROUTING
# ═══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("""
    <div style="padding:20px 10px 10px; text-align:center;">
        <div style="font-size:30px">📋</div>
        <div style="font-size:17px; font-weight:600; color:#e8f0f8; margin:6px 0 2px">Prima Nota Paghe</div>
        <div style="font-size:11px; color:#4a6a8a;">GB Software · Wolters Kluwer</div>
    </div>
    <hr>
    """, unsafe_allow_html=True)

    pagina = st.radio("", ["📋  Buste Paga", "🏦  F24", "⚙️  Regole"],
                      label_visibility="collapsed")

    st.markdown("""
    <hr>
    <div style="font-size:11px; color:#3a5a7a; padding:0 4px; line-height:1.8;">
        <b style="color:#6a9aba;">Come si usa</b><br>
        1. Carica il PDF<br>
        2. Verifica le voci estratte<br>
        3. Scarica XLSX (opzionale)<br>
        4. Genera e scarica XML<br><br>
        Per modificare conti e voci vai su <b style="color:#6a9aba;">Regole</b>.<br>
        Le voci senza regola si compilano a mano senza bisogno di aggiornare nulla.
    </div>
    """, unsafe_allow_html=True)

if "Buste Paga" in pagina:
    pagina_buste_paga()
elif "F24" in pagina:
    pagina_f24()
else:
    pagina_regole()
